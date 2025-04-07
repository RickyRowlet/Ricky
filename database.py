import openpyxl 
import os
from datetime import datetime

FILES = {
    "warehouse": "warehouse.xlsx",
    "sales": "sales.xlsx",
    "accounting": "accounting.xlsx"
}

def create_files():
    for key, file in FILES.items():
        if not os.path.exists(file):
            wb = openpyxl.Workbook()
            ws = wb.active
            if key == "warehouse":
                ws.append(["ID", "Product Name", "Quantity", "Unit Price", "Supplier"])
            elif key == "sales":
                ws.append(["Invoice ID", "Date", "Customer", "Total"])
            elif key == "accounting":
                ws.append(["Transaction ID", "Date", "Type", "Amount", "Description"])
            wb.save(file)

def add_product(name, quantity, price, supplier):
    wb = openpyxl.load_workbook(FILES["warehouse"])
    ws = wb.active
    next_id = 1
    if ws.max_row > 1:
        last_id = ws.cell(row=ws.max_row, column=1).value
        if isinstance(last_id, int):
            next_id = last_id + 1
    ws.append([next_id, name, quantity, price, supplier])
    wb.save(FILES["warehouse"])

def get_products():
    wb = openpyxl.load_workbook(FILES["warehouse"])
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def update_product(product_id, name, quantity, price, supplier):
    wb = openpyxl.load_workbook(FILES["warehouse"])
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[1].value = name
            row[2].value = quantity
            row[3].value = price
            row[4].value = supplier
            break
    wb.save(FILES["warehouse"])

def delete_product(product_id):
    wb = openpyxl.load_workbook(FILES["warehouse"])
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == product_id:
            ws.delete_rows(i)
            break
    wb.save(FILES["warehouse"])

def search_products(keyword):
    all_data = get_products()
    return [row for row in all_data if keyword.lower() in str(row[1]).lower()]

def add_invoice(customer, total_amount):
    wb = openpyxl.load_workbook(FILES["sales"])
    ws = wb.active
    next_id = 1
    if ws.max_row > 1:
        last_id = ws.cell(row=ws.max_row, column=1).value
        if isinstance(last_id, int):
            next_id = last_id + 1
    date = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([next_id, date, customer, total_amount])
    wb.save(FILES["sales"])

def get_invoices():
    wb = openpyxl.load_workbook(FILES["sales"])
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def add_transaction(transaction_type, amount, description):
    wb = openpyxl.load_workbook(FILES["accounting"])
    ws = wb.active
    next_id = 1
    if ws.max_row > 1:
        last_id = ws.cell(row=ws.max_row, column=1).value
        if isinstance(last_id, int):
            next_id = last_id + 1
    date = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([next_id, date, transaction_type, amount, description])
    wb.save(FILES["accounting"])

def get_transactions():
    wb = openpyxl.load_workbook(FILES["accounting"])
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def get_inventory_summary():
    wb = openpyxl.load_workbook(FILES["warehouse"])
    ws = wb.active
    total_quantity = 0
    total_value = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        quantity = row[2]
        unit_price = row[3]
        if isinstance(quantity, (int, float)) and isinstance(unit_price, (int, float)):
            total_quantity += quantity
            total_value += quantity * unit_price
    return total_quantity, total_value

def get_sales_summary():
    wb = openpyxl.load_workbook(FILES["sales"])
    ws = wb.active
    total_sales = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        amount = row[3]
        if isinstance(amount, (int, float)):
            total_sales += amount
    return total_sales

def get_accounting_summary():
    wb = openpyxl.load_workbook(FILES["accounting"])
    ws = wb.active
    total_income = 0
    total_expense = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        trans_type = row[2]
        amount = row[3]
        if isinstance(amount, (int, float)):
            if trans_type.lower() == "thu":
                total_income += amount
            elif trans_type.lower() == "chi":
                total_expense += amount
    return total_income, total_expense
def search_invoices(keyword="", start_date="", end_date=""):
    wb = openpyxl.load_workbook(FILES["sales"])
    ws = wb.active
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        invoice_id, date_str, customer, total = row
        if keyword.lower() in str(customer).lower():
            results.append(row)
        elif start_date and end_date:
            try:
                from datetime import datetime
                invoice_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
                if start <= invoice_date <= end:
                    results.append(row)
            except:
                continue
    return results

def search_transactions(trans_type="", keyword="", start_date="", end_date=""):
    wb = openpyxl.load_workbook(FILES["accounting"])
    ws = wb.active
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        trans_id, date_str, t_type, amount, desc = row
        if trans_type and t_type.lower() != trans_type.lower():
            continue
        if keyword.lower() in str(desc).lower():
            results.append(row)
        elif start_date and end_date:
            try:
                from datetime import datetime
                trans_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
                if start <= trans_date <= end:
                    results.append(row)
            except:
                continue
    return results

def init_user_file():
    file = "users.xlsx"
    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Username", "Password"])
        ws.append(["admin", "admin"])
        wb.save(file)
